"""
Capa de servicios — contiene toda la lógica de negocio.

Responsabilidades:
  - Sesiones de conversación (sesiones_ia)
  - Extracción de datos con OpenAI Function Calling
  - Matching de productos contra el inventario
  - Creación de clientes y cotizaciones en BD
  - Generación de PDF con fpdf2
  - Envío de mensajes y documentos a WhatsApp (Meta API)
"""
from __future__ import annotations

import json
import logging
import os
import re
import subprocess
import zipfile
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

_TZ_MX = ZoneInfo("America/Mexico_City")

import requests
from fpdf import FPDF
from openai import OpenAI
from openpyxl import load_workbook
from sqlalchemy import extract, func
from sqlalchemy.orm import Session

from .config import settings
from .models import Cliente, Cotizacion, DetalleCotizacion, Inventario, Negocio, SesionIA
from .schemas import DatosRenta, ItemEquipo
from data.catalogo import (
    CATALOGO,
    construir_sinonimos,
    detectar_color_en_texto,
    producto_por_nombre,
    tiene_variantes_color,
)

logger = logging.getLogger(__name__)

_openai_client = OpenAI(api_key=settings.OPENAI_API_KEY)

META_API_BASE = "https://graph.facebook.com/v20.0"

# ─── Esquema de función para OpenAI Function Calling ─────────────────────────

_FUNCION_EXTRACCION: dict[str, Any] = {
    "name": "extraer_datos_renta",
    "description": (
        "Extrae la información de renta de equipo para evento a partir del mensaje del cliente. "
        "Solo devuelve los campos que el cliente mencionó explícitamente; omite los demás."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "nombre": {
                "type": "string",
                "description": "Nombre completo del cliente",
            },
            "domicilio": {
                "type": "string",
                "description": (
                    "Solo la calle y número del domicilio de entrega, sin colonia ni ciudad. "
                    "Si el cliente da la dirección en formato 'Calle Num, Colonia, Ciudad', "
                    "extrae aquí solo 'Calle Num'. "
                    "IMPORTANTE: los nombres de vialidades mexicanas pueden incluir preposiciones "
                    "como 'a' (ej. 'Carretera Federal a Tehuacán 469', 'Boulevard a San Andrés 55'). "
                    "En esos casos extrae el nombre COMPLETO de la vialidad incluyendo la preposición "
                    "y el destino, más el número exterior. "
                    "Ejemplos: 'Federal a Tehuacán 469' -> 'Federal a Tehuacán 469'; "
                    "'Blvd Valsequillo 1811' -> 'Blvd Valsequillo 1811'."
                ),
            },
            "colonia": {
                "type": "string",
                "description": (
                    "Colonia, fraccionamiento o barrio de la dirección de entrega. "
                    "Si el cliente escribe 'Blvd Valsequillo 1811, Universidades, Puebla', "
                    "extrae aquí 'Universidades'."
                ),
            },
            "referencia": {
                "type": "string",
                "description": "Referencia o punto de referencia para localizar el domicilio",
            },
            "fecha_entrega": {
                "type": "string",
                "description": "Fecha de entrega del equipo, formato YYYY-MM-DD",
            },
            "hora_entrega": {
                "type": "string",
                "description": (
                    "Hora de entrega del equipo. Puede ser una hora exacta en HH:MM 24h "
                    "(ej. '14:00') O un rango si el cliente lo da (ej. '09:00 - 17:00' "
                    "para '9 a 5', '9 AM a 5 PM', 'entre 9 y 5'). Conserva el rango si el "
                    "cliente lo menciona — no escojas solo una hora."
                ),
            },
            "fecha_evento": {
                "type": "string",
                "description": "Fecha del evento, formato YYYY-MM-DD",
            },
            "fecha_recoleccion": {
                "type": "string",
                "description": (
                    "Fecha de recolección o devolución del equipo, formato YYYY-MM-DD. "
                    "SIEMPRE debe ser >= fecha_evento. "
                    "Si el cliente dice un día de la semana (ej. 'el viernes', 'el sábado'), "
                    "usa la PRIMERA ocurrencia de ese día que sea IGUAL O POSTERIOR a fecha_evento, "
                    "nunca antes. Ejemplo: evento jueves 30 de abril, cliente dice 'el viernes' "
                    "-> 2026-05-01 (el viernes siguiente al evento, NO el viernes anterior)."
                ),
            },
            "maps_link": {
                "type": "string",
                "description": (
                    "URL de Google Maps que el cliente compartió. Reconoce dominios: "
                    "maps.google.com, goo.gl/maps, maps.app.goo.gl, google.com/maps, "
                    "www.google.com.mx/maps. Si el cliente pegó un link tipo "
                    "https://maps.app.goo.gl/abc123 o https://www.google.com/maps/place/..., "
                    "extrae la URL completa aqui."
                ),
            },
            "equipo": {
                "type": "array",
                "description": "Lista de artículos a rentar",
                "items": {
                    "type": "object",
                    "properties": {
                        "cantidad": {"type": "integer", "minimum": 1},
                        "descripcion": {"type": "string"},
                    },
                    "required": ["cantidad", "descripcion"],
                },
            },
            "instrucciones": {
                "type": "string",
                "description": "Instrucciones especiales o notas adicionales del cliente",
            },
            "tipo_entrega": {
                "type": "string",
                "enum": ["domicilio", "recoger"],
                "description": "'domicilio' si el cliente quiere que se lo lleven, 'recoger' si va al local",
            },
            "requiere_factura": {
                "type": "boolean",
                "description": (
                    "True si el cliente indica que quiere/necesita factura "
                    "(palabras como 'si', 'si factura', 'necesito factura', 'con IVA'). "
                    "False si dice 'no', 'sin factura', 'no necesito'."
                ),
            },
            "mantel_color": {
                "type": "string",
                "description": (
                    "Color del mantel que eligió el cliente (ej. 'blanco', 'azul marino', "
                    "'dorado con blanco'). Solo si lo mencionó."
                ),
            },
            "silla_color": {
                "type": "string",
                "description": (
                    "Color de la silla Tiffany. Opciones: 'Blanca', 'Chocolate', 'Plata'. "
                    "Reconoce abreviaturas: 'bca'/'blanco' -> 'Blanca', 'cafe'/'madera' -> 'Chocolate', "
                    "'gris'/'plata' -> 'Plata'. Solo si el cliente mencionó color de silla."
                ),
            },
        },
        "required": [],
    },
}

_MENSAJE_BIENVENIDA = (
    "¡Hola! 👋 Soy el asistente virtual de *Alquiladora Crystal*. "
    "Estoy aquí para ayudarte a cotizar el equipo que necesitas para tu evento. "
    "¿Me puedes decir tu nombre y qué equipo necesitas rentar?"
)

_PREGUNTAS_FALTANTES = {
    "nombre":            "¿A nombre de quién va a quedar la cotización? (nombre completo)",
    "equipo":            "¿Qué equipo necesitas? Por ejemplo: '20 sillas acojinadas y 5 mesas'.",
    "mantel_color":      "¿De qué color necesitas los manteles? Te mando la paleta disponible 👇",
    "silla_color":       "¿De qué color quieres la *Silla Tiffany*? Tenemos: *Blanca*, *Chocolate* o *Plata*.",
    "fecha_evento":      "¿Para qué fecha es el evento? Puedes escribir algo como '26 de abril'.",
    "tipo_entrega":      "¿Prefieres pasar al local a recoger el equipo o te lo llevamos a domicilio?",
    "domicilio": (
        "¿Cuál es tu *dirección*?\n"
        "Dame *calle y número* (ej. _Blvd Valsequillo 1811_).\n\n"
        "💡 *Tip*: también puedes compartirme tu *ubicación de Google Maps*."
    ),
    "colonia":           "¿En qué *colonia o fraccionamiento*? (ej. Universidades, Lomas de Angelópolis)",
    "referencia": (
        "¿Tienes alguna *referencia* para llegar más fácil?\n"
        "Ejemplos: _Cluster 888_, _casa amarilla_, _frente al parque_, _portón negro_."
    ),
    "fechas_grupo": (
        "¡Listo! Ahora las fechas 📅\n\n"
        "📦 *Entrega*: ¿qué día y a qué hora te llegamos con el equipo?\n"
        "🚚 *Recolección*: ¿qué día lo recogemos?\n\n"
        "Ejemplo: _entrega el viernes 24 de abril a las 12 PM, recolección el domingo 26._"
    ),
    "fechas_grupo_recoger": (
        "¡Listo! Ahora las fechas 📅\n\n"
        "📦 *Tú pasas al local*: ¿qué día y a qué hora pasas a recoger el equipo?\n"
        "🚚 *Devolución*: ¿qué día regresas el equipo al local?\n\n"
        "Ejemplo: _paso el viernes 24 de abril a las 12 PM, lo regreso el domingo 26._"
    ),
    "fecha_entrega":     "¿Qué día te entregamos el equipo? (puede ser un día antes del evento)",
    "fecha_entrega_recoger": "¿Qué día pasas al local a recoger el equipo? (puede ser un día antes del evento)",
    "hora_entrega":      "¿A qué hora te llegamos con el equipo?",
    "hora_entrega_recoger": "¿A qué hora pasas al local? (también puedes darme un rango, ej. 'de 9 AM a 5 PM')",
    "fecha_recoleccion": "¿Para qué día programamos la recolección del equipo?",
    "fecha_recoleccion_recoger": "¿Qué día regresas el equipo al local?",
    "requiere_factura":  "¿Vas a necesitar factura? (responde 'si' o 'no' — si es 'si', se agrega el IVA 16%)",
}

_SALUDOS_EXACTOS = {
    "hola", "ola", "holi", "hi", "hey", "hello", "saludos",
    "buenas", "buenos",
    "buenas tardes", "buenos dias", "buenos días",
    "buenas noches", "buen dia", "buen día",
    "buena tarde", "buena noche",
    "hola buenas tardes", "hola buenos dias", "hola buenos días",
    "hola buenas noches", "hola buen dia", "hola buen día",
    "que tal", "qué tal", "k tal", "q tal",
}


def es_saludo(mensaje: str) -> bool:
    """True si el mensaje es un saludo puro sin informacion adicional."""
    t = mensaje.lower().strip().rstrip(".!?¡¿,")
    t = " ".join(t.split())  # normaliza espacios
    return t in _SALUDOS_EXACTOS


# Sinónimos semánticos — se leen del catalogo central en data/catalogo.py
_SINONIMOS_PRODUCTO: dict[str, str] = construir_sinonimos()

# Zonas de flete — clave: palabra clave en domicilio/colonia, valor: precio en pesos
# Ordenadas de más específica a más general para evitar falsos positivos
_ZONAS_FLETE: dict[str, int] = {
    # Zona cercana al negocio (Blvd Valsequillo / Universidades)
    "universidades": 80,
    "valsequillo": 80,
    "anzures": 80,
    "san manuel": 80,
    "xonaca": 80,
    "el carmen": 80,
    "humboldt": 80,
    "tecnologico": 80,
    "tecnológico": 80,
    # Zona media
    "centro": 100,
    "san andres": 120,
    "san andrés": 120,
    "cholula": 130,
    "cuautlancingo": 130,
    # Zona lejana
    "angelopolis": 150,
    "angelópolis": 150,
    "lomas": 150,
    "reserva territorial": 150,
    "chachapa": 250,
    "chachaa": 250,
    "amozoc": 200,
    "tehuacan": 300,
    "tehuacán": 300,
    "atlixco": 280,
}
_FLETE_DEFECTO = 100   # Para colonias no listadas
_FLETE_MINIMO  = 80    # Nunca cobramos menos que esto


# ─── Sesiones ─────────────────────────────────────────────────────────────────

def obtener_o_crear_sesion(db: Session, telefono: str, negocio_id: str) -> SesionIA:
    """Devuelve la sesión activa o crea una nueva."""
    sesion = (
        db.query(SesionIA)
        .filter(
            SesionIA.telefono_cliente == telefono,
            SesionIA.negocio_id == negocio_id,
            SesionIA.activa == True,  # noqa: E712
        )
        .first()
    )
    if not sesion:
        sesion = SesionIA(
            telefono_cliente=telefono,
            negocio_id=negocio_id,
            contexto_actual={},
        )
        db.add(sesion)
        db.commit()
        db.refresh(sesion)
    return sesion


def actualizar_contexto_sesion(db: Session, sesion: SesionIA, nuevo_contexto: dict) -> None:
    sesion.contexto_actual = nuevo_contexto
    sesion.ultimo_mensaje = datetime.utcnow()
    db.commit()


def marcar_sesion_esperando_autorizacion(
    db: Session, sesion: SesionIA, cotizacion_id: str
) -> None:
    """Vincula la cotización a la sesión y la deja activa esperando 'sí autorizo'."""
    sesion.cotizacion_id = cotizacion_id
    sesion.ultimo_mensaje = datetime.utcnow()
    db.commit()


def cerrar_sesion(db: Session, sesion: SesionIA, cotizacion_id: str | None = None) -> None:
    # Elimina sesiones cerradas anteriores del mismo teléfono para evitar el UNIQUE constraint
    db.query(SesionIA).filter(
        SesionIA.telefono_cliente == sesion.telefono_cliente,
        SesionIA.negocio_id == sesion.negocio_id,
        SesionIA.activa == False,  # noqa: E712
    ).delete()
    sesion.activa = False
    if cotizacion_id:
        sesion.cotizacion_id = cotizacion_id
    db.commit()


# ─── Extracción con OpenAI ────────────────────────────────────────────────────

def extraer_datos_ia(
    mensaje: str,
    contexto_previo: dict,
    nombre_negocio: str,
    campo_esperado: str | None = None,
) -> DatosRenta:
    """
    Llama a OpenAI con Function Calling para extraer los datos del mensaje.
    Devuelve solo los campos que el cliente mencionó en este mensaje.
    `campo_esperado` es el campo que el bot acaba de preguntar — sirve para
    desambiguar respuestas cortas ("24 de abril" → fecha_entrega si eso se preguntó).
    """
    contexto_str = json.dumps(contexto_previo, ensure_ascii=False, indent=2) if contexto_previo else "{}"
    ahora_mx = datetime.now(_TZ_MX)
    anio_actual = ahora_mx.year
    fecha_hoy = ahora_mx.strftime("%Y-%m-%d")
    nombre_dia_hoy = _DIAS_ES_CAP[ahora_mx.weekday()]
    fecha_manana = (ahora_mx + timedelta(days=1)).strftime("%Y-%m-%d")

    hint_campo = ""
    if campo_esperado:
        hint_campo = (
            f"\nIMPORTANTE: El bot acaba de preguntarle al cliente por el campo '{campo_esperado}'. "
            f"Si la respuesta del cliente es breve o ambigua (solo una fecha, una hora, un nombre, "
            f"una palabra, un color), asumela como el valor de '{campo_esperado}' y asignala a ese campo. "
            f"No la asignes a un campo que ya tenga valor en los datos recopilados. "
            f"NO vuelvas a extraer 'equipo' a menos que el cliente repita explicitamente la lista "
            f"de productos con cantidades nuevas."
        )

    respuesta = _openai_client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {
                "role": "system",
                "content": (
                    f"Eres el asistente de cotizaciones de {nombre_negocio}, "
                    "una empresa de renta de equipo para eventos en Mexico. "
                    f"\nFECHA Y ZONA HORARIA — HOY ES {nombre_dia_hoy} {fecha_hoy} en Mexico (America/Mexico_City). "
                    f"\n - 'hoy' / 'hoy mismo' -> {fecha_hoy}"
                    f"\n - 'mañana' / 'manana' -> {fecha_manana}"
                    f"\n - 'el viernes', 'el sabado' -> proxima ocurrencia de ese dia desde {fecha_hoy}"
                    f"\n - Si solo dicen '24 de abril' sin año, usa el año {anio_actual}"
                    " a menos que el cliente indique otro año.\n"
                    "IMPORTANTE: Valida que la fecha sea real antes de devolverla. "
                    "Abril, junio, septiembre y noviembre tienen 30 dias. "
                    "Febrero tiene 28 (29 en bisiesto). El resto tiene 31. "
                    "Si el cliente da una fecha imposible (ej. '31 de abril'), omite el campo de fecha. "
                    "FECHAS RELATIVAS: 'al dia siguiente'/'un dia despues' del 30 de abril = 2026-05-01 (1 de mayo). "
                    "Siempre respeta los limites del mes al calcular fechas relativas. "
                    "Para fecha_recoleccion: SIEMPRE debe ser >= fecha_evento. "
                    "Si el cliente dice un dia de la semana ('el viernes', 'el sabado', 'el domingo'), "
                    "calcula la primera ocurrencia de ese dia que sea IGUAL O POSTERIOR a fecha_evento "
                    "(nunca un dia pasado, nunca antes del evento). "
                    "Ejemplo: evento jueves 30-abr, cliente dice 'el viernes' -> 2026-05-01. "
                    "NO uses el viernes anterior al evento aunque sea mas cercano al dia de hoy. "
                    "\n\nREGLAS PARA EXTRAER EQUIPO (MUY IMPORTANTE):\n"
                    "0. Si el cliente dice solo 'mesas' o 'sillas' SIN especificar tipo, "
                    "extrae descripcion EXACTAMENTE como 'mesa' o 'silla'. NUNCA asumas "
                    "'mesa tablon', 'silla acojinada' o cualquier tipo si el cliente no lo dijo. "
                    "El bot le preguntará al cliente qué tipo quiere.\n"
                    "0b. Si en los datos ya recopilados YA HAY items con descripcion generica "
                    "(ej. 'mesa' cantidad 4 o 'silla' cantidad 40) y el cliente AHORA esta "
                    "respondiendo con el tipo (ej. 'mesa tablon', 'silla tiffany blanca'), "
                    "REUSA la cantidad original del item generico. NUNCA pongas cantidad 1 "
                    "por defecto en este caso. Ejemplo: contexto tiene [{4, 'mesa'}, {40, 'silla'}], "
                    "cliente dice 'mesa tablon con manteles' → extrae "
                    "[{4, 'mesa tablon'}, {4, 'mantel tablon'}] (silla queda igual, no se toca).\n"
                    "1. Items compuestos con 'con su/sus' generan MULTIPLES items:\n"
                    "   - '5 tablones con sus manteles' -> "
                    "[{cantidad:5, descripcion:'mesa tablon'}, {cantidad:5, descripcion:'mantel tablon'}]\n"
                    "   - '10 mesas redondas con manteles' -> "
                    "[{cantidad:10, descripcion:'mesa redonda'}, {cantidad:10, descripcion:'mantel redondo'}]\n"
                    "2. Cuando el cliente diga 'para N personas' SIN dar cantidades, infiere asi:\n"
                    "   - sillas, platos, vasos, copas, tenedores, cuchillos, cucharas: N unidades\n"
                    "   - mesas/tablones/manteles: ceil(N / 10) (una mesa cada 10 personas)\n"
                    "   - Ejemplo: 'mesas, sillas y platos para 100 personas' -> "
                    "[{cantidad:10,descripcion:'mesa tablon'}, {cantidad:10,descripcion:'mantel tablon'}, "
                    "{cantidad:100,descripcion:'silla acojinada'}, {cantidad:100,descripcion:'plato'}]\n"
                    "3. Cuando diga 'cubiertos' o 'loza para N personas', incluye: "
                    "N platos + N tenedores + N cuchillos (como items separados).\n"
                    "4. No inventes items que el cliente no mencione. "
                    "Si dice 'sillas y mesas' pero no menciona platos, NO agregues platos.\n"
                    "Cuando mencione una hora como '2 de la tarde' o '14 hrs', conviertela "
                    "a formato HH:MM en 24h."
                    f"\nDatos ya recopilados:\n{contexto_str}"
                    f"{hint_campo}\n"
                    "Extrae SOLO lo que el cliente mencione en su mensaje actual."
                ),
            },
            {"role": "user", "content": mensaje},
        ],
        tools=[{"type": "function", "function": _FUNCION_EXTRACCION}],
        tool_choice={"type": "function", "function": {"name": "extraer_datos_renta"}},
    )

    tool_call = respuesta.choices[0].message.tool_calls[0]
    argumentos = json.loads(tool_call.function.arguments)
    logger.info("OpenAI extrajo: %s", argumentos)

    # Convertir lista de dicts a lista de ItemEquipo
    if "equipo" in argumentos and argumentos["equipo"]:
        argumentos["equipo"] = [ItemEquipo(**item) for item in argumentos["equipo"]]

    # Descartar fechas imposibles que OpenAI devuelva (ej. "2026-04-31")
    for _campo_fecha in ("fecha_entrega", "fecha_evento", "fecha_recoleccion"):
        _val = argumentos.get(_campo_fecha)
        if _val and _parsear_fecha_iso_safe(_val) is None:
            logger.warning("Fecha invalida ignorada: %s = %s", _campo_fecha, _val)
            argumentos.pop(_campo_fecha)

    # Asegurar que fecha_recoleccion sea >= fecha_evento (avanzar 7 días si es antes)
    _recol = argumentos.get("fecha_recoleccion")
    _evt   = argumentos.get("fecha_evento") or contexto_previo.get("fecha_evento")
    if _recol and _evt:
        d_recol = _parsear_fecha_iso_safe(_recol)
        d_evt   = _parsear_fecha_iso_safe(_evt)
        if d_recol and d_evt and d_recol < d_evt:
            corregida = d_recol
            while corregida < d_evt:
                corregida = corregida + timedelta(days=7)
            argumentos["fecha_recoleccion"] = corregida.isoformat()
            logger.info("Recolección corregida: %s → %s (después de evento %s)",
                        _recol, argumentos["fecha_recoleccion"], _evt)

    # Fallback: detectar color de silla directamente del mensaje completo
    # Esto captura "tiffany bca", "sillas tiffany cafe", etc. aunque la IA no lo haya extraido.
    if not argumentos.get("silla_color"):
        variantes = tiene_variantes_color(mensaje)
        if variantes:
            _, colores_dict = variantes
            color = detectar_color_en_texto(mensaje, colores_dict)
            if color:
                argumentos["silla_color"] = color
                logger.info("Color de silla detectado en texto: %s", color)

    # Normalizar silla_color al nombre oficial si vino en forma abreviada
    silla_col = argumentos.get("silla_color")
    if silla_col:
        prod_tiffany = producto_por_nombre("Silla Tiffany")
        if prod_tiffany and "colores" in prod_tiffany:
            if silla_col not in prod_tiffany["colores"]:
                normalizado = detectar_color_en_texto(silla_col, prod_tiffany["colores"])
                if normalizado:
                    argumentos["silla_color"] = normalizado

    # Fallback: detectar si/no para requiere_factura cuando la IA no lo extrajo
    if argumentos.get("requiere_factura") is None and campo_esperado == "requiere_factura":
        factura = _detectar_respuesta_booleana(mensaje)
        if factura is not None:
            argumentos["requiere_factura"] = factura
            logger.info("Factura detectada en texto: %s", factura)

    # Refinar items genéricos: si el cliente acaba de clarificar un tipo
    # (ej. 'mesa' → 'mesa tablon'), heredar cantidades del contexto previo y
    # preservar los items genéricos que aún no se clarificaron.
    _refinar_items_genericos(argumentos, mensaje, contexto_previo)

    # Auto-añadir manteles: cuando el cliente menciona color/manteleria sin cantidad,
    # generamos los manteles que correspondan a las mesas del contexto previo.
    _auto_anadir_manteles(argumentos, mensaje, contexto_previo)

    # Extraer dirección de un link de Google Maps si está presente
    _extraer_info_maps(argumentos, mensaje)

    return DatosRenta.model_validate(argumentos)


_RESPUESTAS_NO = {"no", "nop", "nope", "nel", "negativo", "no factura",
                  "sin factura", "no necesito", "no gracias", "no quiero"}
_RESPUESTAS_SI = {"si", "sí", "yes", "sip", "claro", "si factura",
                  "con factura", "con iva", "necesito factura", "quiero factura"}


def _detectar_respuesta_booleana(mensaje: str) -> bool | None:
    """Detecta 'si'/'no' en respuestas cortas para campos booleanos."""
    t = mensaje.lower().strip().rstrip(".!?¡¿,")
    if t in _RESPUESTAS_NO:
        return False
    if t in _RESPUESTAS_SI:
        return True
    return None


_TIPOS_MESA_KW  = ("tablon", "tablón", "redonda", "redondo", "vintage",
                   "organica", "orgánica", "lunch", "infantil")
_TIPOS_SILLA_KW = ("tiffany", "tifany", "acojinada", "crossback",
                   "americana", "infantil", "plegable", "luis xv")


def _categoria_item(desc: str) -> str | None:
    d = (desc or "").lower()
    if "mantel" in d:
        return "mantel"
    if "mesa" in d:
        return "mesa"
    if "silla" in d:
        return "silla"
    return None


def _es_item_generico(desc: str) -> bool:
    """True si el item no tiene tipo especifico (ej. 'mesa' o 'sillas' a secas)."""
    d = (desc or "").lower().strip()
    cat = _categoria_item(d)
    if cat == "mesa":
        return not any(k in d for k in _TIPOS_MESA_KW)
    if cat == "silla":
        return not any(k in d for k in _TIPOS_SILLA_KW)
    return False


def _desc_de(item: Any) -> str:
    return (item.descripcion if hasattr(item, "descripcion")
            else (item.get("descripcion") if isinstance(item, dict) else "")) or ""


def _cant_de(item: Any) -> int:
    return (item.cantidad if hasattr(item, "cantidad")
            else (item.get("cantidad") if isinstance(item, dict) else 0)) or 0


def _refinar_items_genericos(
    argumentos: dict, mensaje: str, contexto_previo: dict
) -> None:
    """
    Cuando el cliente clarifica el tipo de un item generico ('mesa' → 'mesa tablon'),
    hereda la cantidad original al refinar el item, y preserva los demás items que
    aún no se hayan clarificado (silla generica, etc.).
    """
    nuevos = argumentos.get("equipo")
    if not nuevos:
        return
    previos = contexto_previo.get("equipo") or []
    if not previos:
        return

    # Cantidades de items genericos previos por categoria (mesa/silla)
    cant_generica: dict[str, int] = {}
    items_previos_no_genericos: list[ItemEquipo] = []
    for it in previos:
        desc_p = _desc_de(it)
        cant_p = _cant_de(it)
        if not (desc_p and cant_p):
            continue
        cat = _categoria_item(desc_p)
        if cat in ("mesa", "silla") and _es_item_generico(desc_p):
            cant_generica[cat] = cant_p
        else:
            items_previos_no_genericos.append(
                ItemEquipo(cantidad=cant_p, descripcion=desc_p)
            )

    if not cant_generica:
        return  # nada que refinar

    # Solo heredamos cantidad si el mensaje del cliente NO trae numeros propios
    mensaje_sin_numeros = re.search(r"\d", mensaje) is None

    refinados: list[ItemEquipo] = []
    cats_refinadas: set[str] = set()
    for item in nuevos:
        desc = _desc_de(item)
        cant = _cant_de(item)
        cat = _categoria_item(desc)
        if (
            cat in ("mesa", "silla")
            and cat in cant_generica
            and not _es_item_generico(desc)   # solo refinar si vino con tipo
            and mensaje_sin_numeros
            and cant <= 1
        ):
            cant_heredada = cant_generica[cat]
            logger.info("Cantidad heredada: '%s' → %d (de %s genérico)",
                        desc, cant_heredada, cat)
            cant = cant_heredada
            cats_refinadas.add(cat)
        refinados.append(ItemEquipo(cantidad=cant, descripcion=desc))

    # Categorias previas genericas que NO se refinaron en este mensaje:
    # las preservamos como genericas para que el bot vuelva a preguntar.
    for cat, cant_prev in cant_generica.items():
        if cat not in cats_refinadas:
            refinados.append(ItemEquipo(cantidad=cant_prev, descripcion=cat))

    # Preservar items previos especificos que no estan en lo nuevo
    descs_actuales = {r.descripcion.lower() for r in refinados}
    for it_prev in items_previos_no_genericos:
        if it_prev.descripcion.lower() not in descs_actuales:
            refinados.append(it_prev)

    argumentos["equipo"] = refinados


def _auto_anadir_manteles(argumentos: dict, mensaje: str, contexto_previo: dict) -> None:
    """
    Si el cliente acepta manteles (menciona color o la palabra 'mantel') pero no
    especifica cantidad, generamos los manteles que correspondan a las mesas del
    contexto (refinado o previo).
    """
    msg_lower = mensaje.lower()
    quiere_manteles = (
        argumentos.get("mantel_color")
        or any(k in msg_lower for k in ("mantel", "mantelería", "manteleria"))
    )
    if not quiere_manteles:
        return

    def _tiene_mantel(equipo) -> bool:
        return any("mantel" in _desc_de(it).lower() for it in (equipo or []))

    if _tiene_mantel(argumentos.get("equipo")):
        return
    mesas_previas = contexto_previo.get("equipo") or []
    if _tiene_mantel(mesas_previas):
        return

    # Buscar mesas primero en argumentos.equipo (ya refinado), luego en el contexto previo
    fuente_mesas: list[tuple[str, int]] = []
    for it in (argumentos.get("equipo") or []):
        desc, cant = _desc_de(it), _cant_de(it)
        if "mesa" in desc.lower() and cant:
            fuente_mesas.append((desc, cant))
    if not fuente_mesas:
        for it in mesas_previas:
            desc, cant = _desc_de(it), _cant_de(it)
            if "mesa" in desc.lower() and cant:
                fuente_mesas.append((desc, cant))

    nuevos_manteles: list[ItemEquipo] = []
    for desc, cant in fuente_mesas:
        d = desc.lower()
        if "tablon" in d or "tablón" in d:
            nuevos_manteles.append(ItemEquipo(cantidad=cant, descripcion="mantel tablon"))
        elif "redonda" in d or "redondo" in d:
            nuevos_manteles.append(ItemEquipo(cantidad=cant, descripcion="mantel redondo"))

    if not nuevos_manteles:
        return

    # Asegurar que los items previos que no esten en argumentos.equipo se preserven
    equipo_actual = list(argumentos.get("equipo") or [])
    descs_actuales = {_desc_de(it).lower() for it in equipo_actual}
    for it_prev in mesas_previas:
        desc_p, cant_p = _desc_de(it_prev), _cant_de(it_prev)
        if desc_p and cant_p and desc_p.lower() not in descs_actuales:
            equipo_actual.append(ItemEquipo(cantidad=cant_p, descripcion=desc_p))

    argumentos["equipo"] = equipo_actual + nuevos_manteles
    argumentos["manteleria_consultada"] = True
    logger.info("Manteles auto-añadidos: %s",
                [(m.cantidad, m.descripcion) for m in nuevos_manteles])


_MAPS_URL_RE = re.compile(
    r'https?://(?:maps\.google\.[a-z.]+|goo\.gl/maps|maps\.app\.goo\.gl|www\.google\.[a-z.]+/maps|google\.[a-z.]+/maps)[^\s]+',
    re.IGNORECASE,
)


def _extraer_info_maps(argumentos: dict, mensaje: str) -> None:
    """Si hay un link de Google Maps, lo guarda y extrae lo que pueda (calle, colonia)."""
    url = argumentos.get("maps_link")
    if not url:
        match = _MAPS_URL_RE.search(mensaje)
        if not match:
            return
        url = match.group(0)
        argumentos["maps_link"] = url

    # Resolver short URLs (maps.app.goo.gl) siguiendo redirects
    final_url = url
    try:
        resp = requests.head(url, allow_redirects=True, timeout=5)
        final_url = resp.url
    except requests.RequestException as exc:
        logger.warning("No se pudo resolver maps URL: %s", exc)

    # Buscar /place/<dir>/  → "Calle Numero, Colonia, Ciudad"
    from urllib.parse import unquote
    place_match = re.search(r'/place/([^/@?]+)', final_url)
    if not place_match:
        return
    addr = unquote(place_match.group(1)).replace("+", " ")
    partes = [p.strip() for p in addr.split(",") if p.strip()]
    if partes and not argumentos.get("domicilio"):
        argumentos["domicilio"] = partes[0]
    if len(partes) >= 2 and not argumentos.get("colonia"):
        argumentos["colonia"] = partes[1]
    logger.info("Maps extraído: domicilio=%s, colonia=%s",
                argumentos.get("domicilio"), argumentos.get("colonia"))


def generar_pregunta_faltante(campos: list[str], tipo_entrega: str | None = None) -> str:
    """
    Devuelve la pregunta más natural para el primer campo que falta.
    Si tipo_entrega == 'recoger' usa preguntas adaptadas (cliente pasa por equipo).
    """
    es_recoger = tipo_entrega == "recoger"

    # Caso especial: si faltan TODAS las fechas, preguntar en un solo mensaje
    fechas_grupo = {"fecha_entrega", "hora_entrega", "fecha_recoleccion"}
    if fechas_grupo.issubset(set(campos)):
        previos = {"nombre", "equipo", "fecha_evento", "tipo_entrega", "domicilio",
                   "colonia", "referencia", "silla_color", "mantel_color"}
        if not (previos & set(campos)):
            return _PREGUNTAS_FALTANTES["fechas_grupo_recoger" if es_recoger else "fechas_grupo"]

    orden = [
        "nombre", "equipo", "silla_color", "mantel_color",
        "fecha_evento", "tipo_entrega",
        "domicilio", "colonia", "referencia",
        "fecha_entrega", "hora_entrega", "fecha_recoleccion",
        "requiere_factura",
    ]
    for campo in orden:
        if campo in campos:
            # Variante "_recoger" si aplica y existe
            if es_recoger:
                key_recoger = f"{campo}_recoger"
                if key_recoger in _PREGUNTAS_FALTANTES:
                    return _PREGUNTAS_FALTANTES[key_recoger]
            return _PREGUNTAS_FALTANTES[campo]
    return "¿Hay algún otro detalle que quieras agregar a tu cotización?"


# ─── Inventario ───────────────────────────────────────────────────────────────

# Productos que requieren depósito en garantía (no están en catálogo)
_KEYWORDS_DEPOSITO = {
    "cristaleria", "cristalería", "vaso", "copa", "caliz", "cáliz",
    "loza", "plato", "platon", "platón", "taza", "vajilla",
    "jarra", "charola", "cubierto", "cuchillo", "tenedor",
}

# Categorías para consultas tipo "¿qué sillas manejan?"
_CATEGORIAS_CONSULTA = {
    "silla":    ["silla"],
    "mesa":     ["mesa"],
    "mantel":   ["mantel"],
    "carpa":    ["carpa", "toldo", "lona"],
    "cristaleria": ["cristaleria", "vaso", "copa"],
    "loza":     ["loza", "plato", "vajilla"],
}


# Categorias en el catalogo cuyo nombre contiene la palabra clave
_CATEGORIAS_AMBIGUAS = ("silla", "mesa")
# Palabras de relleno que ignoramos al chequear si la descripcion es generica
_PALABRAS_RELLENO = {"para", "con", "sin", "y", "los", "las", "una", "un", "unos", "unas",
                     "del", "de", "la", "el", "personas", "invitados"}


def _es_descripcion_generica(descripcion: str) -> str | None:
    """
    Si la descripcion solo dice 'silla(s)' o 'mesa(s)' sin modelo, devuelve la categoria.
    Si especifica modelo (tiffany, tablon, redonda, acojinada, etc.), devuelve None.
    """
    palabras = [
        w.lower() for w in descripcion.split()
        if not w.isdigit() and len(w) > 2 and w.lower() not in _PALABRAS_RELLENO
    ]
    palabras_nucleo = [
        w for w in palabras
        if not any(cat in w for cat in _CATEGORIAS_AMBIGUAS)
    ]
    # Si solo queda la palabra de categoria → es generico
    if not palabras_nucleo:
        for cat in _CATEGORIAS_AMBIGUAS:
            if any(cat in w for w in palabras):
                return cat
    return None


_FRASES_DECLINAR_MANTELERIA = (
    "sin mantel", "sin manteleria", "sin mantelería",
    "no mantel", "no manteleria", "no mantelería",
    "no quiero mantel", "no necesito mantel",
)


def manteleria_declinada(mensaje: str) -> bool:
    """True si el mensaje contiene una negacion explicita sobre manteleria."""
    t = mensaje.lower()
    return any(frase in t for frase in _FRASES_DECLINAR_MANTELERIA)


def analizar_equipo_y_clarificar(
    equipo: list,
    _db: Session,
    _negocio_id: str,
    mensaje_usuario: str = "",
    manteleria_consultada: bool = False,
) -> str | None:
    """
    Detecta items genericos (ej. 'sillas') que necesitan especificar modelo,
    y sugiere complementos (manteleria si hay mesas sin manteles).
    Devuelve un mensaje de clarificacion o None si no se requiere.
    """
    if not equipo:
        return None

    genericos: list[tuple[int, str]] = []   # [(cantidad, categoria)]
    tiene_mesa = False
    tiene_mantel = False

    for item in equipo:
        desc = (item.descripcion if hasattr(item, "descripcion") else item.get("descripcion", "")).lower()
        cantidad = item.cantidad if hasattr(item, "cantidad") else item.get("cantidad", 0)

        if any(kw in desc for kw in ("mantel", "cubremantel", "camino")):
            tiene_mantel = True
        if any(kw in desc for kw in ("mesa", "tablon", "tablón")):
            tiene_mesa = True

        cat = _es_descripcion_generica(desc)
        if cat:
            genericos.append((cantidad, cat))

    secciones: list[str] = []

    # Pregunta por modelo cuando la descripcion es generica
    for cantidad, categoria in genericos:
        opciones = [p["nombre"] for p in CATALOGO if categoria in p["nombre"].lower()]
        if len(opciones) > 1:
            secciones.append(
                f"¿Qué tipo de *{categoria}* quieres para tus *{cantidad}*? "
                f"Tenemos: {', '.join(opciones)}."
            )

    # Sugerir manteleria solo si: (a) hay mesa sin mantel, (b) no se ha consultado antes,
    # y (c) el cliente no esta declinando explicitamente en este mensaje.
    saltar_manteleria = manteleria_consultada or manteleria_declinada(mensaje_usuario)
    if tiene_mesa and not tiene_mantel and not saltar_manteleria:
        secciones.append(
            "¿Quieres incluir también *manteles*, *cubremanteles* o *caminos de mesa* para tus mesas? "
            "(Si no, dime 'sin manteleria')."
        )

    if not secciones:
        return None

    return "\n\n".join(secciones) + (
        "\n\n_Por ejemplo: '100 Tiffany blancas y 10 mesa tablón con manteles'._"
    )


_KW_PRECIO = (
    "cuanto cuesta", "cuánto cuesta", "cuanto vale", "cuánto vale",
    "que cuesta", "qué cuesta", "que precio", "qué precio",
    "precio de", "precio del", "precio para", "tarifa de", "tarifa del",
    "costo de", "costo del", "cuanto sale", "cuánto sale",
    "cuanto me cuesta", "cuánto me cuesta",
    "saber el precio", "saber precio",
    "me das precio", "me das el precio",
)


def detectar_consulta_precio(
    texto: str, db: Session, negocio_id: str
) -> Inventario | None:
    """
    Detecta si el cliente esta preguntando solo por el precio de un producto
    (ej. 'cuanto cuesta la silla acojinada', 'que precio tiene la mesa vintage').
    Devuelve el producto si lo encuentra, o None.
    """
    if not texto:
        return None
    t = texto.lower()
    if not any(kw in t for kw in _KW_PRECIO):
        return None
    # Si el mensaje trae cantidades grandes (ej. '100 sillas tiffany'), es mas
    # probable que sea una solicitud de cotizacion, no una consulta de precio.
    if re.search(r"\b(\d{2,})\s*(silla|mesa|mantel|tablon|tablón)", t):
        return None
    return buscar_producto(db, negocio_id, texto)


def responder_consulta_precio(producto: Inventario) -> str:
    """Mensaje breve con el precio + invitacion a cotizar."""
    return (
        f"La *{producto.nombre_producto}* tiene un precio de "
        f"*${producto.precio_renta:,.2f} por unidad*.\n\n"
        "¿Te gustaría que te ayude a cotizar para tu evento? "
        "Si es así, dime tu *nombre* y *cuántas piezas* necesitas. 😊"
    )


def detectar_consulta_inventario(texto: str) -> str | None:
    """
    Detecta preguntas como '¿qué tipo de sillas manejan?' o '¿qué mesas tienen?'.
    Devuelve la categoría preguntada o None si no es una consulta de inventario.
    """
    t = texto.lower()
    patrones_consulta = ["que tipo", "qué tipo", "que sill", "qué sill", "que mesa", "qué mesa",
                         "cuales sill", "cuáles sill", "que modelos", "qué modelos",
                         "tienen sill", "tienen mesa", "manejan sill", "manejan mesa",
                         "que manten", "qué manten", "que carpa", "qué carpa"]
    if not any(p in t for p in patrones_consulta):
        return None
    for categoria, keywords in _CATEGORIAS_CONSULTA.items():
        if any(kw in t for kw in keywords):
            return categoria
    return None


def responder_consulta_inventario(db: Session, negocio_id: str, categoria: str) -> str:
    """Devuelve un mensaje con los productos disponibles de la categoría."""
    productos = (
        db.query(Inventario)
        .filter(
            Inventario.negocio_id == negocio_id,
            Inventario.activo == True,  # noqa: E712
            Inventario.stock_total > 0,
            func.lower(Inventario.nombre_producto).contains(categoria),
        )
        .order_by(Inventario.nombre_producto)
        .limit(20)
        .all()
    )
    if not productos:
        return (
            f"Déjame consultar con el equipo qué {categoria}s tenemos disponibles. "
            "¿Hay algún otro equipo que te interese mientras tanto?"
        )
    lineas = [f"• *{p.nombre_producto}* — ${p.precio_renta:,.0f}" for p in productos]
    return (
        f"Estos son los modelos de {categoria} que manejamos:\n\n"
        + "\n".join(lineas)
        + "\n\n¿Cuál te interesa y cuántas piezas necesitas?"
    )


def _categoria_producto(descripcion: str) -> str | None:
    """Identifica la familia del producto ('silla', 'mesa', etc.) desde la descripción."""
    desc = descripcion.lower()
    for categoria, keywords in _CATEGORIAS_CONSULTA.items():
        if any(kw in desc for kw in keywords):
            return categoria
    return None


def sugerir_alternativas(db: Session, negocio_id: str, descripcion: str) -> list[str]:
    """Devuelve hasta 5 alternativas de la misma categoría del producto no encontrado."""
    categoria = _categoria_producto(descripcion)
    if not categoria:
        return []
    productos = (
        db.query(Inventario)
        .filter(
            Inventario.negocio_id == negocio_id,
            Inventario.activo == True,  # noqa: E712
            Inventario.stock_total > 0,
            func.lower(Inventario.nombre_producto).contains(categoria),
        )
        .order_by(Inventario.precio_renta)
        .limit(5)
        .all()
    )
    return [f"{p.nombre_producto} (${p.precio_renta:,.0f})" for p in productos]


def requiere_deposito(equipo: list) -> bool:
    """True si alguna descripción contiene cristaleria/loza/vajilla."""
    for item in equipo or []:
        desc = (item.descripcion if hasattr(item, "descripcion") else item.get("descripcion", "")).lower()
        if any(kw in desc for kw in _KEYWORDS_DEPOSITO):
            return True
    return False


def buscar_producto(db: Session, negocio_id: str, descripcion: str) -> Inventario | None:
    """
    Busca un producto aplicando primero sinónimos semánticos (con normalización de
    plurales), luego match multi-palabra combinado, y finalmente stem matching.
    """
    desc_lower = descripcion.lower().strip()

    # Normalizar plurales: "mesas" → "mesa", "sillas" → "silla"
    desc_norm = " ".join(
        w[:-1] if w.endswith("s") and len(w) > 3 else w
        for w in desc_lower.split()
    )

    # 1. Sinónimo semántico: busca en la desc original Y en la normalizada
    for sinonimo, nombre_oficial in _SINONIMOS_PRODUCTO.items():
        if sinonimo in desc_lower or sinonimo in desc_norm:
            logger.info("Sinonimo: '%s' → '%s'", descripcion, nombre_oficial)
            # Lookup exacto por nombre oficial — evita falsos positivos de stem
            resultado = (
                db.query(Inventario)
                .filter(
                    Inventario.negocio_id == negocio_id,
                    Inventario.activo == True,  # noqa: E712
                    Inventario.stock_total > 0,
                    func.lower(Inventario.nombre_producto) == nombre_oficial,
                )
                .first()
            )
            if resultado:
                return resultado
            # Si el lookup exacto falla, usar el nombre oficial como base
            desc_norm = nombre_oficial
            break

    palabras = [p for p in desc_norm.split() if len(p) > 3]

    # 2. Match multi-palabra combinado (más específico que word-by-word)
    if len(palabras) > 1:
        query = db.query(Inventario).filter(
            Inventario.negocio_id == negocio_id,
            Inventario.activo == True,  # noqa: E712
            Inventario.stock_total > 0,
        )
        for palabra in palabras:
            query = query.filter(func.lower(Inventario.nombre_producto).contains(palabra))
        resultado = query.first()
        if resultado:
            logger.info("Match multi-palabra: '%s' → '%s'", descripcion, resultado.nombre_producto)
            return resultado

    # 3. Stem matching palabra por palabra (fallback)
    for palabra in palabras:
        patrones = [palabra]
        for recorte in [1, 2, 3]:
            raiz = palabra[:-recorte]
            if len(raiz) >= 4:
                patrones.append(raiz)

        for patron in patrones:
            resultado = (
                db.query(Inventario)
                .filter(
                    Inventario.negocio_id == negocio_id,
                    Inventario.activo == True,  # noqa: E712
                    Inventario.stock_total > 0,
                    func.lower(Inventario.nombre_producto).contains(patron),
                )
                .first()
            )
            if resultado:
                logger.info("Match stem: '%s' → '%s'", descripcion, resultado.nombre_producto)
                return resultado

    logger.warning("Sin match en inventario para: '%s'", descripcion)
    return None


def calcular_flete(domicilio: str, colonia: str = "") -> int:
    """Determina el costo de flete según la zona del domicilio (mínimo $80)."""
    texto = (domicilio + " " + colonia).lower()
    for zona, precio in _ZONAS_FLETE.items():
        if zona in texto:
            logger.info("Flete: zona '%s' detectada → $%s", zona, precio)
            return max(precio, _FLETE_MINIMO)
    logger.info("Flete: zona no reconocida en '%s' → default $%s", texto, _FLETE_DEFECTO)
    return max(_FLETE_DEFECTO, _FLETE_MINIMO)


# ─── Cotizaciones ─────────────────────────────────────────────────────────────

def _generar_folio_cotizacion(db: Session, negocio_id: str) -> str:
    """Folio de cotización: 'COTI00001-26'. Cuenta cotizaciones del año actual."""
    ahora = datetime.now(_TZ_MX)
    yy = str(ahora.year)[-2:]
    count = (
        db.query(Cotizacion)
        .filter(
            Cotizacion.negocio_id == negocio_id,
            Cotizacion.folio_cotizacion.isnot(None),
            extract("year", Cotizacion.created_at) == ahora.year,
        )
        .count()
    )
    return f"COTI{count + 1:05d}-{yy}"


def _generar_folio_pedido(db: Session, negocio_id: str) -> str:
    """Folio de pedido confirmado: '00001-26'. Cuenta solo cotizaciones confirmadas del año."""
    ahora = datetime.now(_TZ_MX)
    yy = str(ahora.year)[-2:]
    count = (
        db.query(Cotizacion)
        .filter(
            Cotizacion.negocio_id == negocio_id,
            Cotizacion.folio_pedido.isnot(None),
            extract("year", Cotizacion.fecha_confirmacion) == ahora.year,
        )
        .count()
    )
    return f"{count + 1:05d}-{yy}"


_PATRONES_AUTORIZACION = (
    "si autorizo", "sí autorizo", "si, autorizo", "sí, autorizo",
    "autorizo", "autorizado",
    "si lo autorizo", "sí lo autorizo",
    "confirmo", "si confirmo", "sí confirmo",
)


def es_autorizacion(texto: str) -> bool:
    """Detecta si el cliente autoriza el pedido tras recibir la cotización."""
    if not texto:
        return False
    t = texto.strip().lower().rstrip(".!¡")
    # Quitar acentos básicos para comparación
    t = t.replace("í", "i").replace("ó", "o").replace("á", "a").replace("é", "e").replace("ú", "u")
    if t in {"si", "sí", "ok", "okay"}:
        # 'sí' a secas no es autorización (ambiguo) — requerimos la palabra "autorizo"
        return False
    return any(p.replace("í", "i").replace("ó", "o") in t for p in _PATRONES_AUTORIZACION)


def confirmar_pedido(db: Session, cotizacion: Cotizacion) -> str:
    """Marca la cotización como confirmada y le asigna folio_pedido. Devuelve el folio."""
    if cotizacion.confirmada and cotizacion.folio_pedido:
        return cotizacion.folio_pedido
    cotizacion.folio_pedido = _generar_folio_pedido(db, str(cotizacion.negocio_id))
    cotizacion.confirmada = True
    cotizacion.fecha_confirmacion = datetime.now(_TZ_MX)
    cotizacion.estatus = "aceptado"
    db.commit()
    db.refresh(cotizacion)
    return cotizacion.folio_pedido


def guardar_cotizacion(
    db: Session,
    telefono: str,
    negocio_id: str,
    datos: DatosRenta,
) -> Cotizacion:
    """
    Crea o actualiza el Cliente y crea la Cotizacion con sus detalles.
    Los artículos sin coincidencia en inventario se guardan en cotizacion.notas.
    """
    # 1. Cliente — upsert por teléfono
    cliente = db.query(Cliente).filter(Cliente.telefono_whatsapp == telefono).first()
    if not cliente:
        cliente = Cliente(telefono_whatsapp=telefono)
        db.add(cliente)

    # Fallback: si la IA metió la colonia dentro del domicilio (formato "Calle Num, Colonia"),
    # la partimos manualmente.
    calle_num, colonia_fallback = _partir_domicilio(datos.domicilio)
    colonia_final = datos.colonia or colonia_fallback

    # Siempre refrescamos nombre/dirección con lo último que dijo el cliente en este chat
    if datos.nombre:
        cliente.nombre = _normalizar_texto(datos.nombre)
    if calle_num:
        cliente.direccion_predeterminada = _normalizar_texto(calle_num)

    db.flush()  # Obtener el ID si es nuevo

    # 2. Construir notas completas con fechas y datos extras
    folio = _generar_folio_cotizacion(db, negocio_id)
    lineas_notas: list[str] = [f"Folio: {folio}"]
    if datos.fecha_entrega:
        lineas_notas.append(f"Fecha entrega: {datos.fecha_entrega}")
    if datos.hora_entrega:
        lineas_notas.append(f"Hora entrega: {datos.hora_entrega}")
    if datos.fecha_recoleccion:
        lineas_notas.append(f"Fecha recoleccion: {datos.fecha_recoleccion}")
    if colonia_final:
        lineas_notas.append(f"Colonia: {_normalizar_texto(colonia_final)}")
    if datos.referencia:
        lineas_notas.append(f"Referencia: {_normalizar_texto(datos.referencia)}")
    if datos.maps_link:
        lineas_notas.append(f"Maps: {datos.maps_link}")
    if datos.mantel_color:
        lineas_notas.append(f"Mantel color: {_normalizar_texto(datos.mantel_color)}")
    if datos.silla_color:
        lineas_notas.append(f"Silla color: {_normalizar_texto(datos.silla_color)}")
    if datos.instrucciones:
        lineas_notas.append(f"Instrucciones: {datos.instrucciones}")

    sin_inventario: list[str] = []

    # 3. Cotizacion header
    fecha_evento_parsed = _parsear_fecha_iso_safe(datos.fecha_evento)
    if datos.fecha_evento and not fecha_evento_parsed:
        lineas_notas.append(f"Fecha evento (revisar): {datos.fecha_evento}")
        logger.warning("Fecha de evento invalida '%s' — se guarda como nota", datos.fecha_evento)

    cotizacion = Cotizacion(
        negocio_id=negocio_id,
        cliente_id=cliente.id,
        fecha_evento=fecha_evento_parsed,
        estatus="borrador",
        folio_cotizacion=folio,
    )
    db.add(cotizacion)
    db.flush()

    # 4. Detalles — match contra inventario
    total = Decimal("0")
    for item in (datos.equipo or []):
        producto = buscar_producto(db, negocio_id, item.descripcion)
        if producto:
            precio = Decimal(str(producto.precio_renta))
            detalle = DetalleCotizacion(
                cotizacion_id=cotizacion.id,
                producto_id=producto.id,
                cantidad=item.cantidad,
                precio_unitario=precio,
            )
            db.add(detalle)
            total += precio * item.cantidad
        else:
            sin_inventario.append(f"{item.cantidad}x {item.descripcion}")

    if sin_inventario:
        lineas_notas.append("Sin precio (verificar): " + ", ".join(sin_inventario))

    # Flete
    flete = Decimal("0")
    if datos.tipo_entrega == "domicilio":
        flete = Decimal(calcular_flete(calle_num or "", colonia_final or ""))
        lineas_notas.append(f"Flete a domicilio: ${flete:,.0f}")
    else:
        lineas_notas.append("Entrega: pasa al local")

    # IVA (solo si el cliente pidió factura)
    iva = Decimal("0")
    if datos.requiere_factura:
        iva = ((total + flete) * Decimal("0.16")).quantize(Decimal("0.01"))
        lineas_notas.append(f"Factura: si (IVA 16% = ${iva:,.2f})")
    else:
        lineas_notas.append("Factura: no")

    # Depósito en garantía (cristaleria/loza/vajilla)
    if requiere_deposito(datos.equipo or []):
        lineas_notas.append("Deposito: requerido por cristaleria/loza (a definir por el equipo)")

    cotizacion.total = total + flete + iva
    cotizacion.notas = "\n".join(lineas_notas) if lineas_notas else None

    db.commit()
    db.refresh(cotizacion)
    return cotizacion


# ─── Generación de PDF — Diseño Crystal ──────────────────────────────────────

# Colores corporativos
_C_AZUL     = (14, 47, 130)    # Azul Crystal (encabezados, líneas)
_C_AZUL_LNK = (0, 112, 192)    # Azul link (email, hora entrega)
_C_NEGRO    = (0, 0, 0)
_C_NARANJA  = (204, 95, 0)     # FOLIO y Depósito
_C_BLANCO   = (255, 255, 255)
# Íconos de redes sociales
_C_IG       = (193, 53, 132)   # Instagram
_C_FB       = (24, 119, 242)   # Facebook
_C_WA       = (37, 211, 102)   # WhatsApp
_C_WEB      = (0, 112, 192)    # Sitio web

# Anchos de columnas de la tabla (total = 190 mm)
_WC = (22, 18, 76, 22, 26, 26)   # Faltante C | Cant | Desc | Faltante Cl | P.Unit | Importe
_TOTAL_TABLA_FILAS = 18           # Filas visibles en la tabla (incluyendo vacías)

_CLAUSULAS = [
    "El cliente reconoce que la firma en cualquier parte de este contrato corresponde a el o a su representante, para efectos legales y mercantiles.",
    "El cliente autoriza ajustar el importe total en caso de modificaciones al pedido original.",
    "El costo del flete se calculara segun la distancia y la cantidad de equipo. Cualquier entrega o recoleccion solicitada fuera del horario laboral (Lunes a Sabado de 9:00 a.m. a 5:30 p.m) incurrira en un costo adicional.",
    "La loza y cristaleria deben devolverse limpias; de lo contrario, se cobrara un 35% adicional sobre la renta.",
    "La entrega y recoleccion se realizara en planta baja y no incluye montaje. En caso contrario, se cobrara un costo adicional por servicio.",
    "El cliente sera responsable del costo de reposicion en caso de perdida parcial o total del equipo, asi como de danos a la manteleria (rasgadura, quemadura, manchas de tinta).",
    "El contrato debe pagarse en su totalidad al momento de la entrega del equipo; de no ser asi, el equipo no podra ser dejado en el domicilio del cliente y se considerara como una rescision del contrato.",
    "El cliente debe proporcionar los datos necesarios de su identidad o del receptor del equipo por razones de seguridad. Nos reservamos el derecho de solicitar informacion o documentos adicionales.",
    "El cliente tiene la obligacion de inspeccionar el equipo al momento de la entrega y notificar cualquier dano o defecto de inmediato. La falta de notificacion puede implicar la aceptacion del equipo en su estado actual.",
    "El cliente debe proteger el equipo de la intemperie; de no hacerlo, se cobrara un 25% adicional sobre el alquiler.",
    "Cualquier cancelacion incurre en una penalizacion del 25% sobre el anticipo.",
]

_PAGARE = (
    "Por este PAGARE me (nos) obligo (amos) a pagar incondicionalmente a la orden de "
    "Jose Luis Castro Conde, en la ciudad de Puebla, Pue., la cantidad de $________ "
    "el dia ___ de __________ 202__. Si a su vencimiento no cubro dicha cantidad, "
    "pagare ademas un interes mensual del ____% asi como los gastos y costos que se "
    "originen, en caso de que el cobro sea de naturaleza judicial. Este pagare se rige "
    "de acuerdo a lo estipulado en el Art. 170 Capitulo III de la Ley General de "
    "Sociedades Mercantiles."
)


def _set_color(pdf: FPDF, rgb: tuple, target: str = "text") -> None:
    r, g, b = rgb
    if target == "text":
        pdf.set_text_color(r, g, b)
    elif target == "fill":
        pdf.set_fill_color(r, g, b)
    else:
        pdf.set_draw_color(r, g, b)


def _linea_campo(
    pdf: FPDF, x: float, y: float, w: float,
    etiqueta: str, valor: str = "", h: float = 5.8, label_w: float = 30,
) -> None:
    """Etiqueta en negrita + línea azul inferior con valor."""
    pdf.set_xy(x, y)
    pdf.set_font("Helvetica", "B", 7.5)
    _set_color(pdf, _C_NEGRO, "text")
    _set_color(pdf, _C_NEGRO, "draw")
    pdf.cell(label_w, h, etiqueta + ":", ln=False)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_xy(x + label_w, y)
    _set_color(pdf, _C_AZUL, "draw")          # subrayado azul
    pdf.cell(w - label_w, h, valor, border="B", ln=False)
    _set_color(pdf, _C_NEGRO, "draw")


def _icono_red(pdf: FPDF, x: float, y: float, color: tuple, simbolo: str) -> None:
    """Círculo coloreado con símbolo blanco para íconos de redes sociales."""
    _set_color(pdf, color, "fill")
    _set_color(pdf, color, "draw")
    pdf.ellipse(x, y, 8, 8, "F")
    _set_color(pdf, _C_BLANCO, "text")
    pdf.set_font("Helvetica", "B", 6)
    pdf.set_xy(x, y + 2)
    pdf.cell(8, 4, simbolo, align="C")
    _set_color(pdf, _C_NEGRO, "text")
    _set_color(pdf, _C_NEGRO, "draw")


_DIAS_ES = ["LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES", "SABADO", "DOMINGO"]
_DIAS_ES_CAP = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]
_MESES_ES = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
             "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]


def _parsear_fecha_iso_safe(iso_date: str | None) -> date | None:
    """
    date.fromisoformat() pero devuelve None si la fecha es invalida
    (ej. '2026-04-31' que no existe, o '31 abril' que no es ISO).
    """
    if not iso_date:
        return None
    try:
        return date.fromisoformat(iso_date)
    except (ValueError, TypeError):
        return None


def _formatear_fecha_larga(iso_date: str | None) -> str:
    """YYYY-MM-DD → 'Viernes 24 de abril de 2026'."""
    if not iso_date:
        return ""
    try:
        dt = datetime.strptime(iso_date, "%Y-%m-%d")
    except ValueError:
        return iso_date
    return f"{_DIAS_ES_CAP[dt.weekday()]} {dt.day} de {_MESES_ES[dt.month - 1]} de {dt.year}"


def _formatear_telefono_cliente(tel: str | None) -> str:
    """5212212664376 → '22-12-66-43-76' (sin lada 52)."""
    if not tel:
        return ""
    if tel.startswith("521") and len(tel) == 13:
        tel = tel[3:]
    elif tel.startswith("52") and len(tel) == 12:
        tel = tel[2:]
    if len(tel) == 10 and tel.isdigit():
        return "-".join(tel[i:i + 2] for i in range(0, 10, 2))
    return tel


def _partir_domicilio(domicilio: str | None) -> tuple[str, str]:
    """
    'Rio Jamapa 118, San Manuel, Puebla' → ('Rio Jamapa 118', 'San Manuel').
    Si no hay coma, devuelve (domicilio, '').
    """
    if not domicilio:
        return "", ""
    partes = [p.strip() for p in domicilio.split(",") if p.strip()]
    if len(partes) >= 2:
        return partes[0], partes[1]
    return partes[0] if partes else "", ""


def _formatear_hora_ampm(hora_hhmm: str | None) -> str:
    """HH:MM (24h) → 'H AM' o 'H PM' (12h). Soporta rangos '09:00 - 17:00' → '9 AM - 5 PM'."""
    if not hora_hhmm:
        return ""
    # Rango: separadores comunes
    for sep in (" - ", "-", " a ", " A "):
        if sep in hora_hhmm:
            partes = hora_hhmm.split(sep, 1)
            return f"{_formatear_hora_ampm(partes[0].strip())} - {_formatear_hora_ampm(partes[1].strip())}"
    try:
        h, _m = hora_hhmm.split(":")
        h = int(h)
    except (ValueError, IndexError):
        return hora_hhmm
    period = "AM" if h < 12 else "PM"
    h12 = h if 1 <= h <= 12 else (h - 12 if h > 12 else 12)
    return f"{h12} {period}"


def _formatear_entrega_crystal(fecha_iso: str | None, hora_hhmm: str | None) -> str:
    """Construye 'VIERNES ANTES DE LAS 12 PM' a partir de fecha YYYY-MM-DD y hora HH:MM."""
    if not fecha_iso:
        return ""
    try:
        dt = datetime.strptime(fecha_iso, "%Y-%m-%d")
        dia = _DIAS_ES[dt.weekday()]
    except ValueError:
        return f"{fecha_iso} {hora_hhmm or ''}".strip()
    hora_txt = _formatear_hora_ampm(hora_hhmm)
    return f"{dia} ANTES DE LAS {hora_txt}" if hora_txt else dia


def _parsear_notas(notas: str | None) -> dict:
    campos: dict = {}
    for linea in (notas or "").splitlines():
        if ": " in linea:
            clave, valor = linea.split(": ", 1)
            campos[clave.strip().lower()] = valor.strip()
    return campos


def _generar_pdf_fpdf(
    cotizacion: Cotizacion,
    cliente: Cliente,
    negocio: Negocio,  # noqa: ARG001 — reservado para multi-tenant
) -> str:
    """Genera la nota de servicio con fpdf2 (fallback si falla LibreOffice)."""

    datos = _parsear_notas(cotizacion.notas)
    folio = cotizacion.folio_cotizacion or datos.get("folio", "") or str(cotizacion.id)[:8].upper()
    hoy = datetime.now(_TZ_MX).strftime("%d/%m/%Y  %H:%M hrs")
    h_fila: float = 4.9

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()
    pdf.set_margins(10, 5, 10)
    _set_color(pdf, _C_NEGRO, "draw")
    pdf.set_line_width(0.2)

    # ═══════════════════════════════════════════════════════
    # 1 · ENCABEZADO
    # ═══════════════════════════════════════════════════════
    logo_path = Path(settings.LOGO_PATH)
    if logo_path.exists():
        pdf.image(str(logo_path), x=10, y=4, w=44)

    pdf.set_font("Helvetica", "", 8)
    _set_color(pdf, _C_NEGRO, "text")
    for yi, txt in enumerate(["Boulevard Valsequillo No. 1811-A", "Col. Universidades, Puebla, Pue."]):
        pdf.set_xy(58, 5 + yi * 4.5)
        pdf.cell(90, 4.5, txt, align="C")

    _set_color(pdf, _C_AZUL_LNK, "text")
    pdf.set_xy(58, 14)
    pdf.cell(90, 4.5, "alquiladoracrystal@hotmail.com", align="C")
    _set_color(pdf, _C_NEGRO, "text")
    pdf.set_xy(58, 18.5)
    pdf.cell(90, 4.5, "Tel. 22-22-33-61-66", align="C")

    # Logo derecho — ancho y alto acotados para no pisar las celdas del encabezado
    logo_small = Path(settings.LOGO_PATH).parent / "logosmall.png"
    if logo_small.exists():
        pdf.image(str(logo_small), x=165, y=5, w=24, h=20)
    elif logo_path.exists():
        pdf.image(str(logo_path), x=165, y=5, w=24, h=20)

    # Línea azul separadora
    _set_color(pdf, _C_AZUL, "draw")
    pdf.set_line_width(0.6)
    pdf.line(10, 30, 200, 30)
    pdf.set_line_width(0.2)
    _set_color(pdf, _C_NEGRO, "draw")

    # ═══════════════════════════════════════════════════════
    # 2 · DATOS DEL CLIENTE (2 columnas)
    # ═══════════════════════════════════════════════════════
    y0 = 32
    xr = 115          # columna derecha
    fecha_evento_str = (
        _formatear_fecha_larga(cotizacion.fecha_evento.isoformat())
        if cotizacion.fecha_evento else ""
    )

    colonia_str    = datos.get("colonia", "")
    referencia_str = datos.get("referencia", "")
    domicilio_base = cliente.direccion_predeterminada or ""
    telefono_display = _formatear_telefono_cliente(cliente.telefono_whatsapp)

    _linea_campo(pdf, 10, y0,      100, "Nombre del Cliente",   cliente.nombre or "",    label_w=34)
    _linea_campo(pdf, xr, y0,       85, "Fecha de Cumpleanos",  "",                      label_w=34)
    _linea_campo(pdf, 10, y0+6,    100, "Domicilio Evento",     domicilio_base,          label_w=34)
    _linea_campo(pdf, xr, y0+6,     85, "Celular",              telefono_display,        label_w=18)
    _linea_campo(pdf, 10, y0+12,   100, "Colonia",              colonia_str,             label_w=18)
    _linea_campo(pdf, xr, y0+12,    85, "Domicilio Cliente",    "",                      label_w=34)
    _linea_campo(pdf, 10, y0+18,   100, "Referencias",          referencia_str,          label_w=24)
    _linea_campo(pdf, xr, y0+18,    85, "Fecha de Elaboracion", hoy,                     label_w=34)

    # ═══════════════════════════════════════════════════════
    # 3 · FECHAS + FOLIO
    # ═══════════════════════════════════════════════════════
    y1 = y0 + 27
    fecha_entrega_larga = _formatear_fecha_larga(datos.get("fecha entrega", ""))
    fecha_recol_larga   = _formatear_fecha_larga(datos.get("fecha recoleccion", ""))

    _linea_campo(pdf, 10, y1,     100, "Fecha de Entrega",    fecha_entrega_larga, label_w=30)
    _linea_campo(pdf, 10, y1+6,   100, "Fecha de Evento",     fecha_evento_str,    label_w=30)
    _linea_campo(pdf, 10, y1+12,  100, "Fecha de Devolucion", fecha_recol_larga,   label_w=30)

    # Caja FOLIO — etiqueta negra + número naranja
    _set_color(pdf, _C_AZUL, "draw")
    pdf.set_line_width(0.8)
    pdf.rect(xr, y1, 85, 20)
    pdf.set_line_width(0.2)
    _set_color(pdf, _C_NEGRO, "draw")

    pdf.set_font("Helvetica", "B", 13)
    _set_color(pdf, _C_NEGRO, "text")
    pdf.set_xy(xr, y1 + 1)
    pdf.cell(85, 7, "FOLIO :", align="C")

    pdf.set_font("Helvetica", "B", 16)
    _set_color(pdf, _C_NARANJA, "text")
    pdf.set_xy(xr, y1 + 10)
    pdf.cell(85, 8, folio, align="C")
    _set_color(pdf, _C_NEGRO, "text")

    # ═══════════════════════════════════════════════════════
    # 4 · TABLA DE ARTÍCULOS
    # ═══════════════════════════════════════════════════════
    y2 = y1 + 23
    encabezados = ["Faltante Crystal", "Cantidad", "Descripcion", "Faltante Cliente", "P. Unitario", "Importe"]
    alineaciones = ["C", "C", "L", "C", "R", "R"]

    pdf.set_xy(10, y2)
    _set_color(pdf, _C_AZUL, "fill")
    _set_color(pdf, _C_BLANCO, "text")
    _set_color(pdf, _C_AZUL, "draw")
    pdf.set_font("Helvetica", "B", 7.5)
    for enc, w in zip(encabezados, _WC):
        pdf.cell(w, 7, enc, border=1, align="C", fill=True)

    _set_color(pdf, _C_NEGRO, "draw")
    _set_color(pdf, _C_NEGRO, "text")
    pdf.set_font("Helvetica", "", 8)
    y_fila = y2 + 7
    subtotal_renta = Decimal("0")

    mantel_color_str = datos.get("mantel color", "")
    silla_color_str  = datos.get("silla color", "")

    for i in range(_TOTAL_TABLA_FILAS):
        pdf.set_xy(10, y_fila)
        if i < len(cotizacion.detalles):
            det = cotizacion.detalles[i]
            sub = Decimal(str(det.precio_unitario)) * det.cantidad
            subtotal_renta += sub
            nombre_prod = det.producto.nombre_producto
            nombre_lower = nombre_prod.lower()
            if "mantel" in nombre_lower and mantel_color_str:
                nombre_prod = f"{nombre_prod} - {mantel_color_str}"
            if "tiffany" in nombre_lower and silla_color_str:
                nombre_prod = f"{nombre_prod} - {silla_color_str}"
            valores = ["", str(det.cantidad), nombre_prod, "",
                       f"${det.precio_unitario:,.2f}", f"${sub:,.2f}"]
        else:
            valores = ["", "", "", "", "", ""]
        for val, w, al in zip(valores, _WC, alineaciones):
            pdf.cell(w, h_fila, val, border=1, align=al)
        y_fila += h_fila

    # ═══════════════════════════════════════════════════════
    # 5 · TOTALES (derecha) + INFO ENTREGA (izquierda)
    # ═══════════════════════════════════════════════════════
    y3 = y_fila
    flete_raw = datos.get("flete a domicilio", "0").replace("$", "").replace(",", "")
    flete_val    = Decimal(flete_raw or "0")

    factura_txt  = datos.get("factura", "no").lower()
    con_factura  = factura_txt.startswith("si")
    iva_val      = ((subtotal_renta + flete_val) * Decimal("0.16")).quantize(Decimal("0.01")) if con_factura else Decimal("0")

    requiere_dep = "deposito" in datos  # clave generada en guardar_cotizacion
    deposito_val = Decimal("0")         # monto a definir por el equipo
    anticipo_val = Decimal("0")
    total_val    = subtotal_renta + flete_val + iva_val
    resta_val    = total_val - anticipo_val

    x_tot = 10 + sum(_WC[:4])
    w_et, w_val = _WC[4], _WC[5]

    def _fila_total(etiq: str, valor: Decimal, y: float, bold: bool = False,
                    color_et: tuple = _C_NEGRO, texto_val: str | None = None) -> None:
        pdf.set_xy(x_tot, y)
        pdf.set_font("Helvetica", "B" if bold else "", 7.5)
        _set_color(pdf, color_et, "text")
        pdf.cell(w_et, h_fila, etiq, border=1, align="R")
        _set_color(pdf, _C_NEGRO, "text")
        pdf.set_font("Helvetica", "B" if bold else "", 7.5)
        valor_str = texto_val if texto_val is not None else f"${valor:,.2f}"
        pdf.cell(w_val, h_fila, valor_str, border=1, align="R")

    _fila_total("Subtotal",      subtotal_renta, y3)
    _fila_total("Flete",         flete_val,      y3 + h_fila)
    _fila_total("I.V.A",         iva_val,        y3 + h_fila * 2)
    _fila_total(
        "Deposito", deposito_val, y3 + h_fila * 3, color_et=_C_NARANJA,
        texto_val="POR DEFINIR" if requiere_dep else "$0.00",
    )
    _fila_total("Total a Pagar", total_val,      y3 + h_fila * 4, bold=True)
    _fila_total("Anticipo",      anticipo_val,   y3 + h_fila * 5)
    _fila_total("Resta a Pagar", resta_val,      y3 + h_fila * 6, bold=True)

    # Entrega (izquierda) — el repartidor llena a mano entrega/recoleccion
    fecha_entrega_iso = datos.get("fecha entrega", "")
    hora_entrega_str  = datos.get("hora entrega", "")
    entrega_display   = _formatear_entrega_crystal(fecha_entrega_iso, hora_entrega_str)

    pdf.set_font("Helvetica", "B", 7.5)
    _set_color(pdf, _C_NEGRO, "text")
    pdf.set_xy(10, y3 + 1)
    pdf.cell(38, 5, "Dia y Hora de Entrega:", ln=False)
    if entrega_display:
        _set_color(pdf, _C_AZUL, "text")
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(60, 5, entrega_display.upper(), align="C")
    _set_color(pdf, _C_NEGRO, "text")

    # "Recoger el dia ___ a partir de las ___" — lineas vacias para llenar a mano
    pdf.set_xy(10, y3 + 8)
    pdf.set_font("Helvetica", "", 7.5)
    pdf.cell(18, 5, "Recoger el dia", ln=False)
    _set_color(pdf, _C_AZUL, "draw")
    pdf.cell(26, 5, "", border="B", ln=False)
    _set_color(pdf, _C_NEGRO, "draw")
    pdf.cell(20, 5, "  a partir de las", ln=False)
    _set_color(pdf, _C_AZUL, "draw")
    pdf.cell(24, 5, "", border="B", ln=False)
    _set_color(pdf, _C_NEGRO, "draw")

    # "Entrega ___ Recoleccion ___" — lineas vacias para llenar a mano
    pdf.set_xy(10, y3 + 14)
    pdf.cell(14, 5, "Entrega:", ln=False)
    _set_color(pdf, _C_AZUL, "draw")
    pdf.cell(28, 5, "", border="B", ln=False)
    _set_color(pdf, _C_NEGRO, "draw")
    pdf.cell(22, 5, "  Recoleccion:", ln=False)
    _set_color(pdf, _C_AZUL, "draw")
    pdf.cell(24, 5, "", border="B", ln=False)
    _set_color(pdf, _C_NEGRO, "draw")

    # Fila Anticipos / Pagado
    pdf.set_xy(10, y3 + 20)
    pdf.set_font("Helvetica", "", 7.5)
    pdf.cell(18, 5, "Anticipos", ln=False)
    _set_color(pdf, _C_AZUL, "draw")
    pdf.cell(30, 5, "", border="B", ln=False)
    _set_color(pdf, _C_NEGRO, "draw")
    pdf.cell(12, 5, "   Pagado", ln=False)
    _set_color(pdf, _C_AZUL, "draw")
    pdf.cell(28, 5, "", border="B", ln=False)
    _set_color(pdf, _C_NEGRO, "draw")

    # ═══════════════════════════════════════════════════════
    # 6 · CLÁUSULAS + REDES SOCIALES
    # ═══════════════════════════════════════════════════════
    y4 = y3 + h_fila * 7 + 4
    _set_color(pdf, _C_AZUL, "draw")
    pdf.set_line_width(0.4)
    pdf.line(10, y4, 200, y4)
    pdf.set_line_width(0.2)
    _set_color(pdf, _C_NEGRO, "draw")
    y4 += 1.5

    # "C L A U S U L A S" vertical — con acento en A
    letras_vertical = list("CLAUSULAS")
    letras_vertical[2] = "A"  # la A con acento — Latin-1 OK
    pdf.set_font("Helvetica", "B", 7.5)
    _set_color(pdf, _C_AZUL, "text")
    for idx, letra in enumerate(letras_vertical):
        pdf.set_xy(10, y4 + idx * 5.4)
        pdf.cell(5, 5.4, letra, align="C")
    _set_color(pdf, _C_NEGRO, "text")

    # Texto de cláusulas (4, 10, 11 en negrita)
    x_cl, y_cl = 16, y4
    clausulas_negrita = {4, 10, 11}
    for i, texto in enumerate(_CLAUSULAS, 1):
        pdf.set_xy(x_cl, y_cl)
        es_bold = i in clausulas_negrita
        pdf.set_font("Helvetica", "B", 6.5)
        pdf.cell(5, 4, f"{i}.", ln=False)
        pdf.set_font("Helvetica", "B" if es_bold else "", 6.5)
        pdf.set_xy(x_cl + 5, y_cl)
        pdf.multi_cell(120, 4, texto)
        y_cl = pdf.get_y() + 0.5

    # Redes sociales (columna derecha)
    x_red = 143
    y_red = y4 + 1
    pdf.set_font("Helvetica", "B", 8)
    _set_color(pdf, _C_NEGRO, "text")
    pdf.set_xy(x_red, y_red)
    pdf.cell(57, 5, "Datos y clausulas leidas:", align="C")
    y_red += 8

    redes = [
        (_C_IG,  "IG",  "alquiladora.crystal"),
        (_C_FB,  "f",   "Alquiladora y Banquetes Crystal"),
        (_C_WA,  "WA",  "2226503588"),
        (_C_WEB, "www", "alquiladoracrystal.com"),
    ]
    for color_ic, simbolo, texto_red in redes:
        _icono_red(pdf, x_red, y_red, color_ic, simbolo)
        _set_color(pdf, _C_NEGRO, "text")
        pdf.set_font("Helvetica", "", 7.5)
        pdf.set_xy(x_red + 10, y_red + 1.5)
        pdf.cell(47, 5, texto_red)
        y_red += 10

    # ═══════════════════════════════════════════════════════
    # 7 · PAGARÉ
    # ═══════════════════════════════════════════════════════
    y5 = max(y_cl + 1, y4 + 56)
    pdf.set_xy(10, y5)
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(0, 5, "Pagare", ln=True)
    pdf.set_xy(10, pdf.get_y())
    pdf.set_font("Helvetica", "", 7)
    pdf.multi_cell(190, 4, _PAGARE)

    # Firma
    y6 = min(pdf.get_y() + 3, 290)
    _set_color(pdf, _C_NEGRO, "draw")
    pdf.set_xy(140, y6)
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(60, 6, "Nombre y Firma", border="T", align="C")

    # Guardar
    os.makedirs(settings.PDF_DIR, exist_ok=True)
    nombre_archivo = f"cotizacion_{folio}.pdf"
    ruta = os.path.join(settings.PDF_DIR, nombre_archivo)
    pdf.output(ruta)
    return ruta


# ─── Generación de PDF vía Excel + LibreOffice ───────────────────────────────

def generar_pdf_cotizacion(
    cotizacion: Cotizacion,
    cliente: Cliente,
    negocio: Negocio,  # noqa: ARG001
) -> str:
    """
    Llena la plantilla Excel (assets/Nota.xlsx) con los datos de la cotización
    y la convierte a PDF usando LibreOffice headless.
    Si falla cualquier paso, cae al generador fpdf2 como respaldo.
    """
    template_path = Path(settings.EXCEL_TEMPLATE_PATH)
    soffice_path  = Path(settings.LIBREOFFICE_PATH)

    if not template_path.exists() or not soffice_path.exists():
        logger.warning(
            "Plantilla Excel o LibreOffice no disponible — usando fpdf2. "
            "template=%s  soffice=%s", template_path, soffice_path,
        )
        return _generar_pdf_fpdf(cotizacion, cliente, negocio)

    try:
        return _generar_pdf_excel(cotizacion, cliente, template_path, soffice_path)
    except Exception as exc:
        logger.exception("Error generando PDF con Excel: %s — usando fpdf2", exc)
        return _generar_pdf_fpdf(cotizacion, cliente, negocio)


def _generar_pdf_excel(
    cotizacion: Cotizacion,
    cliente: Cliente,
    template_path: Path,
    soffice_path: Path,
) -> str:
    datos  = _parsear_notas(cotizacion.notas)
    folio  = cotizacion.folio_cotizacion or datos.get("folio", "") or str(cotizacion.id)[:8].upper()
    hoy    = datetime.now(_TZ_MX).strftime("%d/%m/%Y  %H:%M hrs")

    import warnings
    from openpyxl.styles import PatternFill
    _BLANCO = PatternFill(fill_type="solid", fgColor="FFFFFFFF")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")   # suprime el aviso de WMF — las restauramos luego
        wb = load_workbook(template_path)
    ws = wb.active

    def _up(txt: str | None) -> str:
        return (txt or "").upper()

    def _set(coord: str, value) -> None:
        """Escribe valor y asegura fondo blanco (evita gris por defecto de LibreOffice)."""
        cell = ws[coord]
        cell.value = value
        if not cell.fill or cell.fill.fill_type in (None, "none"):
            cell.fill = _BLANCO

    def _set_rc(row: int, col: int, value) -> None:
        cell = ws.cell(row=row, column=col, value=value)
        if not cell.fill or cell.fill.fill_type in (None, "none"):
            cell.fill = _BLANCO

    # ── Datos del cliente ──────────────────────────────────────────────────────
    _set("B9",  _up(_normalizar_texto(cliente.nombre)))
    _set("B10", _up(_normalizar_texto(cliente.direccion_predeterminada)))
    _set("B11", _up(_normalizar_texto(datos.get("colonia", ""))))
    _set("B12", _up(_normalizar_texto(datos.get("referencia", ""))))
    _set("I10", _formatear_telefono_cliente(cliente.telefono_whatsapp))
    _set("I12", hoy.upper())

    # ── Fechas y Folio ─────────────────────────────────────────────────────────
    _set("B15", _formatear_fecha_larga(datos.get("fecha entrega", "")).upper())
    _set("B16", _formatear_fecha_larga(
        cotizacion.fecha_evento.isoformat() if cotizacion.fecha_evento else ""
    ).upper())
    _set("B17", _formatear_fecha_larga(datos.get("fecha recoleccion", "")).upper())
    _set("I15", folio)

    # ── Tabla de artículos (filas 21–41 en Excel, max 21 items) ───────────────
    mantel_color = datos.get("mantel color", "")
    silla_color  = datos.get("silla color", "")
    subtotal = Decimal("0")

    for i, det in enumerate(cotizacion.detalles):
        if i >= 21:
            break
        fila = 21 + i
        sub  = Decimal(str(det.precio_unitario)) * det.cantidad
        subtotal += sub

        nombre_prod  = det.producto.nombre_producto.upper()
        if "MANTEL" in nombre_prod and mantel_color:
            nombre_prod = f"{nombre_prod} - {mantel_color.upper()}"
        if "TIFFANY" in nombre_prod and silla_color:
            nombre_prod = f"{nombre_prod} - {silla_color.upper()}"

        _set_rc(fila, 2,  det.cantidad)
        _set_rc(fila, 3,  nombre_prod)
        _set_rc(fila, 9,  float(det.precio_unitario))
        _set_rc(fila, 10, float(sub))

    # ── Totales ────────────────────────────────────────────────────────────────
    flete_raw   = datos.get("flete a domicilio", "0").replace("$", "").replace(",", "")
    flete_val   = Decimal(flete_raw or "0")
    con_factura = datos.get("factura", "no").lower().startswith("si")
    iva_val     = (
        ((subtotal + flete_val) * Decimal("0.16")).quantize(Decimal("0.01"))
        if con_factura else Decimal("0")
    )
    total_val = subtotal + flete_val + iva_val

    _set("J42", float(subtotal))
    _set("J43", float(flete_val))
    _set("J44", float(iva_val))
    # J45 = Depósito — el equipo lo define a mano
    _set("J46", float(total_val))
    # J47 = Descuento — vacío
    _set("J48", 0.0)
    _set("J49", float(total_val))   # Resta = Total − Anticipo (anticipo=0)

    # ── Info de entrega (solo la celda del repartidor B43) ─────────────────────
    entrega_display = _formatear_entrega_crystal(
        datos.get("fecha entrega", ""), datos.get("hora entrega", "")
    )
    _set("B43", entrega_display.upper())
    # B46 ("Recoger el día") y B48 ("Entrega/Recolección") las llena el chofer a mano

    # ── Guardar .xlsx ─────────────────────────────────────────────────────────
    os.makedirs(settings.PDF_DIR, exist_ok=True)
    xlsx_path = os.path.join(settings.PDF_DIR, f"nota_{folio}.xlsx")
    wb.save(xlsx_path)

    # ── Restaurar imágenes WMF del template (openpyxl las descarta al cargar) ──
    _restaurar_imagenes_xlsx(str(template_path), xlsx_path)

    # ── Convertir a PDF con LibreOffice ───────────────────────────────────────
    result = subprocess.run(
        [str(soffice_path), "--headless", "--convert-to", "pdf",
         "--outdir", settings.PDF_DIR, xlsx_path],
        capture_output=True, text=True, timeout=60,
    )
    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice falló: {result.stderr}")

    pdf_path = xlsx_path.replace(".xlsx", ".pdf")
    if not Path(pdf_path).exists():
        raise FileNotFoundError(f"PDF no generado: {pdf_path}")

    logger.info("PDF generado vía Excel: %s", pdf_path)
    return pdf_path


def _restaurar_imagenes_xlsx(template_path: str, output_path: str) -> None:
    """
    openpyxl descarta imágenes WMF al cargar/guardar.
    Restauramos todos los archivos de media, drawings y relaciones
    copiándolos del template original al output generado.
    """
    _PREFIJOS = ("xl/media/", "xl/drawings/", "xl/worksheets/_rels/")
    try:
        archivos = {}
        with zipfile.ZipFile(template_path, "r") as zin:
            for name in zin.namelist():
                if any(name.startswith(p) for p in _PREFIJOS):
                    archivos[name] = zin.read(name)
        if not archivos:
            return
        tmp = output_path + ".tmp"
        with zipfile.ZipFile(output_path, "r") as zsrc, \
             zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zdst:
            for name in zsrc.namelist():
                data = archivos.get(name) or zsrc.read(name)
                zdst.writestr(name, data)
            for name, data in archivos.items():
                if name not in zsrc.namelist():
                    zdst.writestr(name, data)
        os.replace(tmp, output_path)
        logger.info("Imágenes WMF restauradas desde template (%d archivos)", len(archivos))
    except Exception as exc:
        logger.warning("No se pudieron restaurar imágenes: %s", exc)


# ─── WhatsApp API (Meta) ───────────────────────────────────────────────────────

def _normalizar_texto(texto: str | None) -> str | None:
    """Convierte a Title Case y limpia espacios extra. Acepta None."""
    if not texto:
        return texto
    return " ".join(texto.strip().title().split())


def _normalizar_telefono_mx(telefono: str) -> str:
    """
    WhatsApp envía números mexicanos como 521XXXXXXXXXX (13 dígitos).
    Meta espera 52XXXXXXXXXX (12 dígitos) para enviar mensajes.
    """
    if telefono.startswith("521") and len(telefono) == 13:
        return "52" + telefono[3:]
    return telefono


def _headers_meta() -> dict[str, str]:
    return {
        "Authorization": f"Bearer {settings.WHATSAPP_TOKEN}",
        "Content-Type": "application/json",
    }


def enviar_imagen_whatsapp(telefono: str, ruta_imagen: str, caption: str = "") -> bool:
    """Sube una imagen local a Meta y la envia al cliente con caption opcional."""
    if not Path(ruta_imagen).exists():
        logger.warning("Imagen no encontrada: %s", ruta_imagen)
        return False

    # 1. Subir imagen
    url_upload = f"{META_API_BASE}/{settings.WHATSAPP_PHONE_NUMBER_ID}/media"
    headers_upload = {"Authorization": f"Bearer {settings.WHATSAPP_TOKEN}"}
    mime = "image/png" if ruta_imagen.lower().endswith(".png") else "image/jpeg"
    try:
        with open(ruta_imagen, "rb") as f:
            resp = requests.post(
                url_upload,
                headers=headers_upload,
                files={"file": (Path(ruta_imagen).name, f, mime)},
                data={"messaging_product": "whatsapp", "type": mime},
                timeout=30,
            )
        resp.raise_for_status()
        media_id = resp.json().get("id")
    except requests.RequestException as exc:
        logger.error("Error subiendo imagen a Meta: %s", exc)
        return False

    if not media_id:
        return False

    # 2. Enviar imagen
    url_send = f"{META_API_BASE}/{settings.WHATSAPP_PHONE_NUMBER_ID}/messages"
    payload = {
        "messaging_product": "whatsapp",
        "to": _normalizar_telefono_mx(telefono),
        "type": "image",
        "image": {"id": media_id, "caption": caption},
    }
    try:
        resp = requests.post(url_send, json=payload, headers=_headers_meta(), timeout=10)
        if not resp.ok:
            logger.error("Error enviando imagen WhatsApp a %s: %s — %s", telefono, resp.status_code, resp.text)
            return False
        return True
    except requests.RequestException as exc:
        logger.error("Error enviando imagen WhatsApp: %s", exc)
        return False


def enviar_mensaje_texto(telefono: str, mensaje: str) -> bool:
    """Envía un mensaje de texto al número de WhatsApp dado."""
    url = f"{META_API_BASE}/{settings.WHATSAPP_PHONE_NUMBER_ID}/messages"
    payload = {
        "messaging_product": "whatsapp",
        "to": _normalizar_telefono_mx(telefono),
        "type": "text",
        "text": {"body": mensaje},
    }
    try:
        resp = requests.post(url, json=payload, headers=_headers_meta(), timeout=10)
        if not resp.ok:
            logger.error("Error enviando mensaje WhatsApp a %s: %s — %s", telefono, resp.status_code, resp.text)
            return False
        return True
    except requests.RequestException as exc:
        logger.error("Error enviando mensaje WhatsApp a %s: %s", telefono, exc)
        return False


def _subir_media_whatsapp(ruta_pdf: str) -> str | None:
    """Sube el PDF a Meta y devuelve el media_id."""
    url = f"{META_API_BASE}/{settings.WHATSAPP_PHONE_NUMBER_ID}/media"
    headers = {"Authorization": f"Bearer {settings.WHATSAPP_TOKEN}"}
    try:
        with open(ruta_pdf, "rb") as f:
            resp = requests.post(
                url,
                headers=headers,
                files={"file": (Path(ruta_pdf).name, f, "application/pdf")},

                data={"messaging_product": "whatsapp", "type": "application/pdf"},
                timeout=30,
            )
        resp.raise_for_status()
        return resp.json().get("id")
    except requests.RequestException as exc:
        logger.error("Error subiendo PDF a Meta: %s", exc)
        return None


def enviar_documento_pdf(telefono: str, ruta_pdf: str, caption: str = "") -> bool:
    """Sube el PDF y lo envía como documento de WhatsApp."""
    media_id = _subir_media_whatsapp(ruta_pdf)
    if not media_id:
        return False

    url = f"{META_API_BASE}/{settings.WHATSAPP_PHONE_NUMBER_ID}/messages"
    payload = {
        "messaging_product": "whatsapp",
        "to": _normalizar_telefono_mx(telefono),
        "type": "document",
        "document": {
            "id": media_id,
            "caption": caption,
            "filename": Path(ruta_pdf).name,
        },
    }
    try:
        resp = requests.post(url, json=payload, headers=_headers_meta(), timeout=10)
        resp.raise_for_status()
        return True
    except requests.RequestException as exc:
        logger.error("Error enviando documento WhatsApp a %s: %s", telefono, exc)
        return False
